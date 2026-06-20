import json
import os
import logging
import threading
from django.shortcuts import render, get_object_or_404, redirect
from django.urls import reverse
from django.http import HttpResponse, Http404, JsonResponse
from django.conf import settings
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.contrib.auth import login, authenticate
from .forms import StatementUploadForm, RegisterForm, LoginForm
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

from .models import StatementUpload
from .statement_parser import parse_statement_pdf, is_pdf_encrypted

logger = logging.getLogger(__name__)


def home(request):
    return render(request, 'converter/home.html')


def register(request):
    if request.user.is_authenticated:
        return redirect('home')
    if request.method == 'POST':
        form = RegisterForm(request.POST)
        if form.is_valid():
            user = form.save()
            login(request, user)
            if request.headers.get('x-requested-with') == 'XMLHttpRequest':
                return JsonResponse({'success': True, 'redirect': reverse('upload')})
            messages.success(request, 'Account created successfully!')
            return redirect('upload')
        else:
            if request.headers.get('x-requested-with') == 'XMLHttpRequest':
                return JsonResponse({'success': False, 'errors': form.errors}, status=400)
    else:
        form = RegisterForm()
    return render(request, 'registration/register.html', {'form': form})


@login_required
def profile(request):
    total_conversions = StatementUpload.objects.filter(user=request.user, processed=True, excel_file__isnull=False).count()
    return render(request, 'converter/profile.html', {'total_conversions': total_conversions})


COLUMN_WIDTHS = {
    'A': 16,
    'B': 50,
    'C': 18,
    'D': 18,
    'E': 18,
}


def _style_excel(ws):
    header_font = Font(name='Calibri', bold=True, color='FFFFFF', size=11)
    header_fill = PatternFill(start_color='2F5496', end_color='2F5496', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin'),
    )

    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    for col_letter, width in COLUMN_WIDTHS.items():
        ws.column_dimensions[col_letter].width = width

    data_font = Font(name='Calibri', size=10)
    data_alignment = Alignment(vertical='center', wrap_text=True)
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=5):
        for cell in row:
            cell.font = data_font
            cell.alignment = data_alignment
            cell.border = thin_border

    ws.row_dimensions[1].height = 25


def _process_pdf_background(upload_id, bank, password=None):
    try:
        upload = StatementUpload.objects.get(pk=upload_id)
        pdf_path = upload.pdf_file.path
        logger.info(f"Background processing PDF ({bank}): {pdf_path}")

        df = parse_statement_pdf(pdf_path, bank, password=password)

        if df.empty:
            logger.warning(f"No data extracted for upload {upload_id}")
            upload.processing = False
            upload.save()
            return

        excel_filename = f"{upload.filename()}_converted.xlsx"
        excel_path = os.path.join(settings.MEDIA_ROOT, 'exports', excel_filename)
        os.makedirs(os.path.dirname(excel_path), exist_ok=True)

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Statement'
        headers = ['Date', 'Particulars', 'Withdrawn amount', 'Deposit amount', 'Balance']
        ws.append(headers)
        for _, row in df.iterrows():
            ws.append([
                row.get('Date', ''),
                row.get('Particulars', ''),
                row.get('Withdrawn amount', ''),
                row.get('Deposit amount', ''),
                row.get('Balance', ''),
            ])
        _style_excel(ws)
        wb.save(excel_path)

        upload.excel_file.name = f'exports/{excel_filename}'
        upload.row_count = len(df)

        raw_preview = df.head(15).to_dict('records')
        preview_rows = []
        for r in raw_preview:
            preview_rows.append({
                'd': r.get('Date', ''),
                'p': r.get('Particulars', ''),
                'w': r.get('Withdrawn amount', 0),
                'dep': r.get('Deposit amount', 0),
                'b': r.get('Balance', 0),
            })
        upload.preview_json = json.dumps(preview_rows)
        upload.processed = True
        upload.processing = False
        upload.save()

        logger.info(f"Background processing complete for upload {upload_id}: {len(df)} rows")

    except Exception as e:
        logger.exception(f"Background processing failed for upload {upload_id}: {e}")
        try:
            upload = StatementUpload.objects.get(pk=upload_id)
            upload.processing = False
            upload.save()
        except Exception:
            pass


@login_required
def upload_statement(request):
    needs_password = False
    existing_upload_id = None

    if request.method == 'POST':
        upload_id = request.POST.get('upload_id')

        if upload_id:
            upload = get_object_or_404(StatementUpload, pk=upload_id, user=request.user)
            bank = request.POST.get('bank_name', 'auto')
            password = request.POST.get('pdf_password', '') or None

            upload.processing = True
            upload.save()

            thread = threading.Thread(
                target=_process_pdf_background,
                args=(upload.pk, bank),
                kwargs={'password': password},
                daemon=True
            )
            thread.start()

            return render(request, 'converter/upload.html', {
                'form': StatementUploadForm(),
                'processing': True,
                'upload_id': upload.pk,
            })

        form = StatementUploadForm(request.POST, request.FILES)
        if form.is_valid():
            bank = form.cleaned_data.get('bank_name', 'auto')
            password = form.cleaned_data.get('pdf_password', '') or None

            upload = StatementUpload(user=request.user, pdf_file=form.cleaned_data['pdf_file'])
            upload.save()

            pdf_path = upload.pdf_file.path
            if password is None and is_pdf_encrypted(pdf_path):
                form.add_error('pdf_password', 'This PDF is password-protected. Enter the password to unlock it.')
                return render(request, 'converter/upload.html', {
                    'form': form,
                    'needs_password': True,
                    'existing_upload_id': upload.pk,
                })

            upload.processing = True
            upload.save()

            thread = threading.Thread(
                target=_process_pdf_background,
                args=(upload.pk, bank),
                kwargs={'password': password},
                daemon=True
            )
            thread.start()

            return render(request, 'converter/upload.html', {
                'form': StatementUploadForm(),
                'processing': True,
                'upload_id': upload.pk,
            })
    else:
        form = StatementUploadForm()

    return render(request, 'converter/upload.html', {
        'form': form,
        'needs_password': needs_password,
        'existing_upload_id': existing_upload_id,
    })


@login_required
def check_status(request, pk):
    upload = get_object_or_404(StatementUpload, pk=pk)
    data = {
        'processing': upload.processing,
        'processed': upload.processed,
        'has_excel': bool(upload.excel_file),
        'row_count': upload.row_count,
        'preview_rows': upload.get_preview_rows(),
    }
    if upload.processed and upload.excel_file:
        data['download_url'] = reverse('download', args=[pk])
    return JsonResponse(data)


@login_required
def download_excel(request, pk):
    upload = get_object_or_404(StatementUpload, pk=pk)
    if not upload.excel_file:
        raise Http404('Excel file not found.')
    file_path = upload.excel_file.path
    if not os.path.exists(file_path):
        raise Http404('Excel file not found on disk.')
    with open(file_path, 'rb') as f:
        response = HttpResponse(f.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="{os.path.basename(file_path)}"'
        return response
